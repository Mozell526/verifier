"""In-memory coverage-gate overlay vs already-disqualified name machines.

Does not import or patch draft/judge.py.
Does not treat mixed-pack `role` as an exit switch.
Negative controls (wide / surname / role) are copied only to compare;
they are not candidates.
"""
from __future__ import annotations

import inspect
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
NEW_XLSX = Path("/Users/xiaozijian/Downloads/verifier-client_search-cases-20260814-185013.xlsx")
PACK = Path(__file__).with_name("name_scenario_mixed_pack.json")
RUNS = Path(__file__).with_name("name_scenario_runs")
SET_B = ROOT / "impl/projects/client_search/draft/cases/head_set_b.json"
OUT = Path(__file__).with_name("simulate_1a_coverage_program.json")

BARE_NAME = re.compile(r"^[\u4e00-\u9fff]{2,4}$")
KINSHIP = re.compile(r"[爸妈父母夫妻儿女弟兄姐姊妹叔伯姑舅甥]")
ACTION_FAMILY = re.compile(r"盘客|圈客")
BUSINESS_NON_NAME = re.compile(r"险$|保单|生存金|分红|贷款|口令|团体|意健|外国人|权益")
SEGMENT_OBJECT = re.compile(r"(客户|续收|财富分群)$")
PERSON_THEN_POLICY = re.compile(r"^([\u4e00-\u9fff]{2,4})(?:的)?保单号?$")
LAST_YEAR = re.compile(r"去年|投保日")
PARKED_QUERY = re.compile(r"弟弟|老板娘|大写P07|C00OO|配$")

SURNAMES = set(
    "赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜"
    "戚谢邹喻柏水窦章云苏潘葛奚范彭郎鲁韦昌马苗凤花方俞任袁柳酆鲍史唐"
    "费廉岑薛雷贺倪汤滕殷罗毕郝邬安常乐于时傅皮卞齐康伍余元卜顾孟平黄"
    "和穆萧尹姚邵湛汪祁毛禹狄米贝明臧计伏成戴谈宋茅庞熊纪舒屈项祝董梁"
    "杜阮蓝闵席季麻强贾路娄危江童颜郭梅盛林刁钟徐邱骆高夏蔡田樊胡凌霍"
    "虞万支柯昝管卢莫经房裘缪干解应宗丁宣贲邓郁单杭洪包诸左石崔吉钮龚"
    "程嵇邢滑裴陆荣翁荀羊於惠甄曲家封芮羿储靳汲邴糜松井段富巫乌焦巴弓"
    "牧隗山谷车侯宓蓬全郗班仰秋仲伊宫宁仇栾暴甘钭厉戎祖武符刘景詹束龙"
    "叶幸司韶郜黎蓟薄印宿白怀蒲邰从鄂索咸籍赖卓蔺屠蒙池乔阴鬱胥能苍双"
    "闻莘党翟谭贡劳逄姬申扶堵冉宰郦雍卻璩桑桂濮牛寿通边扈燕冀郏浦尚农"
    "温别庄晏柴瞿阎充慕连茹习宦艾鱼容向古易慎戈廖庾终暨居衡步都耿满弘"
    "匡国文寇广禄阙东欧殳沃利蔚越夔隆师巩厍聂晁勾敖融冷訾辛阚那简饶空"
    "曾毋沙乜养鞠须丰巢关蒯相查后荆红游竺权逯盖益桓公"
)
COMPOUND = (
    "欧阳", "太史", "端木", "上官", "司马", "东方", "独孤", "南宫",
    "万俟", "闻人", "夏侯", "诸葛", "尉迟", "公孙", "慕容", "长孙",
    "宇文", "司徒", "轩辕", "令狐",
)
NAME_FIELDS = {"searchClientName"}
ID_FIELDS = {"clientNo", "polNo"}
CATALOG_FIELDS = {
    "polNoInfo.plancodeinfo.abbrname",
    "polNoInfo.plancodeinfo.plantypedesc",
    "pCategorys",
    "contactOrHomeAddress",
    "pajjMemberGradeInfo.pajjmemberstatus",
    "pajjMemberGradeInfo.pajjmemberproductname",
    "clientNo",
    "polNo",
}


def parse_query(raw: object) -> str:
    if isinstance(raw, str) and raw.strip().startswith("{"):
        try:
            data = json.loads(raw)
            return str(data.get("user_text") or data.get("query") or "").strip()
        except json.JSONDecodeError:
            pass
    match = re.search(r'"user_text"\s*:\s*"([^"]+)"', str(raw or ""))
    return match.group(1) if match else ""


def parse_conditions(raw: object) -> list[dict]:
    if not isinstance(raw, str) or not raw.strip().startswith("{"):
        return []
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return []
    conds = data.get("conditions")
    return [item for item in conds if isinstance(item, dict)] if isinstance(conds, list) else []


def has_surname_shape(token: str) -> bool:
    if not token or not re.fullmatch(r"[\u4e00-\u9fff]{2,4}", token):
        return False
    if any(token.startswith(item) for item in COMPOUND):
        return True
    return token[0] in SURNAMES


def is_parked(row: dict) -> bool:
    query = row["query"]
    if LAST_YEAR.search(query):
        return True
    if PARKED_QUERY.search(query):
        return True
    if KINSHIP.search(query) and "searchClientName" in row["fields"]:
        return True
    return False


def is_action_family(query: str) -> bool:
    return bool(ACTION_FAMILY.search(query) or query == "潜客")


def live_catalog_only(row: dict) -> bool:
    fields = set(row.get("fields") or [])
    return bool(fields & CATALOG_FIELDS) and not (fields & NAME_FIELDS)


def catalog_product_in_query(query: str, catalog: dict[str, set[str]]) -> bool:
    return any(product and product in query for product in catalog["products"])


def whole_query_cover(row: dict) -> str | None:
    """One live field whose value is exactly the whole query. Field-agnostic."""
    fields = list(row.get("fields") or [])
    values = [str(value).strip() for value in (row.get("values") or [])]
    if len(fields) != 1 or len(values) != 1:
        return None
    if values[0] != row["query"]:
        return None
    return fields[0]


def load_set_a() -> list[dict]:
    workbook = load_workbook(NEW_XLSX, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, raw))
        query = parse_query(rec.get("Input / Live Request"))
        conds = parse_conditions(rec.get("Output / 被评估输出"))
        fields = [str(item.get("field") or "") for item in conds if item.get("field")]
        values = []
        for item in conds:
            value = item.get("value")
            if isinstance(value, list):
                values.extend(str(part) for part in value)
            elif value is not None:
                values.append(str(value))
        rows.append(
            {
                "id": rec.get("ID"),
                "query": query,
                "current": rec.get("状态"),
                "fields": fields,
                "values": values,
                "source": "set_a_xlsx",
            }
        )
    return rows


def build_catalog(set_a: list[dict]) -> dict[str, set[str]]:
    products: set[str] = set()
    addresses: set[str] = set()
    for row in set_a:
        for field, value in zip(row["fields"], row["values"] or []):
            token = str(value).strip()
            if not token:
                continue
            if field == "polNoInfo.plancodeinfo.abbrname":
                products.add(token)
            if field == "contactOrHomeAddress":
                addresses.add(token)
    if "金凤" in products:
        products.add("金风")
    return {"products": products, "addresses": addresses}


def load_collected() -> dict[str, dict]:
    out = {}
    if not RUNS.exists():
        return out
    for path in RUNS.glob("*.json"):
        rec = json.loads(path.read_text(encoding="utf-8"))
        out[rec["id"]] = rec
    return out


def load_mixed_rows(set_a: list[dict], collected: dict[str, dict]) -> list[dict]:
    pack = json.loads(PACK.read_text(encoding="utf-8"))
    xlsx = {row["id"]: row for row in set_a}
    rows = []
    for item in pack["cases"]:
        live = ((collected.get(item["id"]) or {}).get("live") or {})
        xrow = xlsx.get(item["id"])
        fields = list(live.get("fields") or (xrow["fields"] if xrow else []))
        values = list(live.get("values") or (xrow["values"] if xrow else []))
        fresh = collected.get(item["id"]) or {}
        fresh_status = fresh.get("judge_status")
        if fresh_status in {"fulfilled", "not_fulfilled"}:
            current = fresh_status
            current_source = "fresh_judge"
        elif xrow:
            current = xrow["current"]
            current_source = "xlsx_judge_fallback" if fresh_status else "xlsx_judge"
        else:
            current = None
            current_source = f"fresh_{fresh_status}" if fresh_status else "untested"
        rows.append(
            {
                "id": item["id"],
                "query": item["query"],
                "pack_role": item["role"],
                "expected": item.get("expected_status"),
                "source": item["source"],
                "fields": fields,
                "values": values,
                "current": current,
                "current_source": current_source,
                "judge_error": fresh.get("judge_error"),
                "live_error": fresh.get("live_error"),
            }
        )
    return rows


def exit_wide(row: dict, catalog: dict[str, set[str]]) -> str | None:
    query = row["query"]
    if is_parked(row) or row.get("pack_role") == "undecided_given_name":
        return None
    if is_action_family(query):
        return "not_fulfilled"
    if live_catalog_only(row):
        return None
    if BARE_NAME.fullmatch(query):
        return "fulfilled"
    return None


def exit_surname(row: dict, catalog: dict[str, set[str]]) -> str | None:
    query = row["query"]
    if is_parked(row) or row.get("pack_role") == "undecided_given_name":
        return None
    if is_action_family(query):
        return "not_fulfilled"
    if live_catalog_only(row):
        return None
    if BUSINESS_NON_NAME.search(query):
        if "searchClientName" in row["fields"]:
            return "not_fulfilled"
        return None
    if catalog_product_in_query(query, catalog) and "searchClientName" in row["fields"]:
        return "not_fulfilled"
    if BARE_NAME.fullmatch(query):
        if has_surname_shape(query):
            return "fulfilled"
        if "searchClientName" in row["fields"]:
            return "not_fulfilled"
    return None


def exit_role(row: dict, catalog: dict[str, set[str]]) -> str | None:
    """Rejected name-type machine. Kept only as a negative control."""
    query = row["query"]
    if is_parked(row) or row.get("pack_role") == "undecided_given_name":
        return None
    if is_action_family(query):
        return None
    if live_catalog_only(row):
        return None
    if SEGMENT_OBJECT.search(query):
        return None
    matched = PERSON_THEN_POLICY.fullmatch(query)
    if matched:
        person = matched.group(1)
        if (
            has_surname_shape(person)
            and "searchClientName" in row["fields"]
            and person in (row.get("values") or [])
        ):
            return "fulfilled"
        return None
    if catalog_product_in_query(query, catalog) and "searchClientName" in row["fields"]:
        return "not_fulfilled"
    if BARE_NAME.fullmatch(query):
        if has_surname_shape(query) and "searchClientName" in row["fields"]:
            return "fulfilled"
        return None
    return None


def exit_live_identity(row: dict, catalog: dict[str, set[str]]) -> str | None:
    """Coverage-gate: overlay only when live covered the whole query as one identity.

    Same function for person names and legal IDs. Does not read pack_role.
    Does not parse 保单号 / 的 / 买了. Surname table is catalog input for 1A,
    not a question-type router.
    """
    field = whole_query_cover(row)
    if field is None:
        return None
    query = row["query"]
    if field in ID_FIELDS:
        return "fulfilled"
    if field != "searchClientName":
        return None
    if not BARE_NAME.fullmatch(query):
        return None
    if query in catalog["products"] or query in catalog["addresses"]:
        return None
    if has_surname_shape(query):
        return "fulfilled"
    return None


def apply_exit(row: dict, exit_status: str | None) -> tuple[str | None, str]:
    if exit_status is None:
        return row.get("current"), "inherit"
    return exit_status, "overlay"


def score_pack(rows: list[dict], key: str) -> dict:
    labeled = [row for row in rows if row.get("expected") in {"fulfilled", "not_fulfilled"}]
    tested = [row for row in labeled if row.get(key) in {"fulfilled", "not_fulfilled"}]
    untested = [row["id"] for row in labeled if row.get(key) not in {"fulfilled", "not_fulfilled"}]
    agree = [row for row in tested if row[key] == row["expected"]]
    miss = [
        {
            "id": row["id"],
            "query": row["query"],
            "pack_role": row["pack_role"],
            "expected": row["expected"],
            "got": row.get(key),
            "mode": row.get(f"{key}_mode"),
            "fields": row.get("fields"),
        }
        for row in tested
        if row[key] != row["expected"]
    ]
    by_role = {}
    for role in sorted({row["pack_role"] for row in labeled}):
        items = [row for row in labeled if row["pack_role"] == role]
        role_tested = [row for row in items if row.get(key) in {"fulfilled", "not_fulfilled"}]
        role_agree = [row for row in role_tested if row[key] == row["expected"]]
        by_role[role] = {
            "n": len(items),
            "tested": len(role_tested),
            "agree": len(role_agree),
        }
    return {
        "n_labeled": len(labeled),
        "tested": len(tested),
        "agree": len(agree),
        "disagree": len(miss),
        "untested": untested,
        "misses": miss,
        "by_role": by_role,
        "status_counts": dict(Counter(row.get(key) for row in rows)),
    }


def mode_report(rows: list[dict], key: str) -> dict:
    overlay = [row for row in rows if row.get(f"{key}_mode") == "overlay"]
    inherit = [row for row in rows if row.get(f"{key}_mode") == "inherit"]
    return {
        "overlay_n": len(overlay),
        "inherit_n": len(inherit),
        "overlay_ids": [row["id"] for row in overlay],
        "overlay_rows": [
            {
                "id": row["id"],
                "query": row["query"],
                "pack_role": row.get("pack_role"),
                "current": row.get("current"),
                "got": row.get(key),
                "expected": row.get("expected"),
                "fields": row.get("fields"),
                "values": row.get("values"),
            }
            for row in overlay
        ],
    }


def live_facts(rows: list[dict]) -> dict:
    set_b = [row for row in rows if str(row["id"]).startswith("HB")]
    name_plus = [row for row in set_b if row["pack_role"] == "name_plus_product"]
    dropped = [
        {
            "id": row["id"],
            "query": row["query"],
            "fields": row["fields"],
            "values": row["values"],
        }
        for row in name_plus
        if "searchClientName" not in row["fields"]
    ]
    return {
        "set_b_n": len(set_b),
        "bare_name_live_ok": [
            row["id"]
            for row in set_b
            if row["pack_role"] == "true_bare_name" and "searchClientName" in row["fields"]
        ],
        "name_plus_product_name_dropped": dropped,
        "name_plus_product_empty": [row["id"] for row in name_plus if not row["fields"]],
        "legal_id_live_ok": [
            row["id"]
            for row in set_b
            if row["pack_role"] == "legal_id" and set(row["fields"]) & ID_FIELDS
        ],
    }


def gate_set_a(rows: list[dict], key: str) -> dict:
    probes = {
        "panke_or_circle": [row for row in rows if is_action_family(row["query"])],
        "fake_name": [row for row in rows if row["query"] in {"共展", "豆芽", "见光", "傻生"}],
        "head_name": [row for row in rows if row["query"] in {"杨杰", "郑鑫", "匡西永", "王坤林", "昊轩"}],
        "catalog_product": [
            row
            for row in rows
            if row["query"] in {"金凤", "宝贝卡", "孝心", "满意", "陇佑智盛"}
        ],
    }
    report = {}
    for name, items in probes.items():
        report[name] = [
            {
                "id": row["id"],
                "query": row["query"],
                "status": row.get(key),
                "mode": row.get(f"{key}_mode"),
            }
            for row in items
        ]
    return report


def set_a_flips(rows: list[dict], key: str) -> dict:
    lifted = []
    dropped = []
    for row in rows:
        before = row.get("current")
        after = row.get(key)
        if before == after:
            continue
        item = {
            "id": row["id"],
            "query": row["query"],
            "before": before,
            "after": after,
            "fields": row.get("fields"),
            "values": row.get("values"),
        }
        if before == "not_fulfilled" and after == "fulfilled":
            lifted.append(item)
        elif before == "fulfilled" and after == "not_fulfilled":
            dropped.append(item)
    return {"lifted_to_f": lifted, "dropped_to_nf": dropped}


def business_cells(rows: list[dict], key: str) -> dict:
    wanted = {
        "yangjie_wangkunlin_same_f": ["杨杰", "王坤林"],
        "fake_still_nf": ["共展", "豆芽"],
        "haoxuan_abstain_or_inherit": ["昊轩"],
        "catalog_jinfeng_keep": ["金凤"],
        "panke_keep_nf": ["盘客"],
    }
    out = {}
    by_query = {row["query"]: row for row in rows}
    for name, queries in wanted.items():
        out[name] = [
            {
                "id": by_query[query]["id"],
                "query": query,
                "status": by_query[query].get(key),
                "mode": by_query[query].get(f"{key}_mode"),
                "current": by_query[query].get("current"),
            }
            for query in queries
            if query in by_query
        ]
    return out


def source_len() -> dict:
    return {
        "exit_wide": len(inspect.getsource(exit_wide).splitlines()),
        "exit_surname": len(inspect.getsource(exit_surname).splitlines()),
        "exit_role": len(inspect.getsource(exit_role).splitlines()),
        "exit_live_identity": len(inspect.getsource(exit_live_identity).splitlines()),
        "whole_query_cover": len(inspect.getsource(whole_query_cover).splitlines()),
    }


def annotate(rows: list[dict], catalog: dict[str, set[str]], exits: dict) -> None:
    for row in rows:
        for name, fn in exits.items():
            status, mode = apply_exit(row, fn(row, catalog))
            row[name] = status
            row[f"{name}_mode"] = mode


def main() -> None:
    set_a = load_set_a()
    catalog = build_catalog(set_a)
    collected = load_collected()
    mixed = load_mixed_rows(set_a, collected)
    exits = {
        "wide": exit_wide,
        "surname": exit_surname,
        "role": exit_role,
        "live_identity": exit_live_identity,
    }
    annotate(set_a, catalog, exits)
    annotate(mixed, catalog, exits)

    mixed_scores = {name: score_pack(mixed, name) for name in ["current", *exits]}
    payload = {
        "note": (
            "Coverage-gate experiment. current=fresh judge or xlsx. "
            "wide/surname/role are negative controls. live_identity is the only candidate. "
            "41/47 is not a KPI. Name+product live drops are facts, not overlay scores."
        ),
        "source_lines": source_len(),
        "catalog_projection": {
            "n_products": len(catalog["products"]),
            "products": sorted(catalog["products"]),
        },
        "live_facts": live_facts(mixed),
        "set_a": {
            "n": len(set_a),
            "current": dict(Counter(row["current"] for row in set_a)),
            **{name: dict(Counter(row[name] for row in set_a)) for name in exits},
            "gates": {name: gate_set_a(set_a, name) for name in ["current", *exits]},
            "flips": {name: set_a_flips(set_a, name) for name in exits},
            "modes": {name: mode_report(set_a, name) for name in exits},
        },
        "mixed": {
            "n": len(mixed),
            "current_sources": dict(Counter(row["current_source"] for row in mixed)),
            "scores": mixed_scores,
            "modes": {name: mode_report(mixed, name) for name in exits},
            "business_cells": {name: business_cells(mixed, name) for name in ["current", *exits]},
            "rows": [
                {
                    "id": row["id"],
                    "query": row["query"],
                    "pack_role": row["pack_role"],
                    "expected": row["expected"],
                    "fields": row["fields"],
                    "values": row["values"],
                    "current": row["current"],
                    "current_source": row["current_source"],
                    "wide": row["wide"],
                    "wide_mode": row["wide_mode"],
                    "surname": row["surname"],
                    "surname_mode": row["surname_mode"],
                    "role": row["role"],
                    "role_mode": row["role_mode"],
                    "live_identity": row["live_identity"],
                    "live_identity_mode": row["live_identity_mode"],
                }
                for row in mixed
            ],
        },
        "set_b_file": {
            "path": str(SET_B),
            "n": len(json.loads(SET_B.read_text(encoding="utf-8"))["cases"]),
        },
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    def compact_score(val: dict) -> dict:
        return {
            "tested": val["tested"],
            "agree": val["agree"],
            "disagree": val["disagree"],
            "untested": val["untested"],
            "misses": [item["id"] + ":" + str(item["got"]) + "/" + str(item.get("mode")) for item in val["misses"]],
        }

    print(json.dumps({
        "wrote": str(OUT),
        "source_lines": payload["source_lines"],
        "live_facts": {
            "set_b_n": payload["live_facts"]["set_b_n"],
            "bare_name_live_ok": payload["live_facts"]["bare_name_live_ok"],
            "name_plus_product_dropped_ids": [
                item["id"] for item in payload["live_facts"]["name_plus_product_name_dropped"]
            ],
            "legal_id_live_ok": payload["live_facts"]["legal_id_live_ok"],
        },
        "set_a_counts": {
            "current": payload["set_a"]["current"],
            **{name: payload["set_a"][name] for name in exits},
        },
        "set_a_flip_n": {
            name: {
                "lifted": len(payload["set_a"]["flips"][name]["lifted_to_f"]),
                "dropped": len(payload["set_a"]["flips"][name]["dropped_to_nf"]),
                "lifted_ids": [item["id"] + ":" + item["query"] for item in payload["set_a"]["flips"][name]["lifted_to_f"]],
                "dropped_ids": [item["id"] + ":" + item["query"] for item in payload["set_a"]["flips"][name]["dropped_to_nf"]],
            }
            for name in exits
        },
        "mixed_scores": {name: compact_score(mixed_scores[name]) for name in mixed_scores},
        "mixed_modes": {
            name: {
                "overlay_n": payload["mixed"]["modes"][name]["overlay_n"],
                "inherit_n": payload["mixed"]["modes"][name]["inherit_n"],
                "overlay_ids": payload["mixed"]["modes"][name]["overlay_ids"],
            }
            for name in exits
        },
        "mixed_business": payload["mixed"]["business_cells"]["live_identity"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
