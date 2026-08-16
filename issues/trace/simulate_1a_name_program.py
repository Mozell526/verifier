"""In-memory overlay of the 1A name exit. Does not import or patch judge.py."""
from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
NEW_XLSX = Path("/Users/xiaozijian/Downloads/verifier-client_search-cases-20260814-185013.xlsx")
SET_B = ROOT / "impl/projects/client_search/draft/cases/head_set_b.json"
OUT = Path(__file__).with_name("simulate_1a_name_program.json")

BARE_NAME = re.compile(r"^[\u4e00-\u9fff]{2,4}$")
ID_TOKEN = re.compile(r"[CPCAcp][A-Za-z0-9]{8,}")
KINSHIP = re.compile(r"[爸妈父母夫妻儿女弟兄姐姊妹叔伯姑舅甥]")
PANKE = re.compile(r"盘客|圈客|潜客")
BUSINESS_NON_NAME = re.compile(
    r"险$|保单|生存金|分红|贷款|口令|团体|意健|外国人|权益"
)
PRODUCT_HINT = re.compile(r"重疾|年金|两全|医疗险|增额寿|增额")
LAST_YEAR = re.compile(r"去年|投保日")
# parked: 称谓 / 格式外空条件 / 弟弟年龄. Overlay must abstain.
PARKED_QUERY = re.compile(r"弟弟|老板娘|大写P07|C00OO|配$")

SURNAMES = set(
    "赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜"
    "戚谢邹喻柏水窦章云苏潘葛奚范彭郎鲁韦昌马苗凤花方俞任袁柳酆鲍史唐"
    "费廉岑薛雷贺倪汤滕殷罗毕郝邬安常乐于时傅皮卞齐康伍余元卜顾孟平黄"
    "和穆萧尹姚邵湛汪祁毛禹狄米贝明臧计伏成戴谈宋茅庞熊纪舒屈项祝董梁"
    "杜阮蓝闵席季麻强贾路娄危江童颜郭梅盛林刁钟徐邱骆高夏蔡田樊胡凌霍"
    "虞万支柯昝管卢莫经房裘缪干解应宗丁宣贲邓郁单杭洪包诸左石崔吉钮龚"
    "程嵇邢滑裴陆荣翁荀羊於惠甄曲家封芮羿储靳汲邴糜松井段富巫乌焦巴弓"
    "牧隗山谷车侯宓蓬全郗班仰秋仲伊宫宁仇栾暴甘钭厉戎祖武符刘景詹束龙"
    "叶幸司韶郜黎蓟薄印宿白怀蒲邰从鄂索咸籍赖卓蔺屠蒙池乔阴鬱胥能苍双"
    "闻莘党翟谭贡劳逄姬申扶堵冉宰郦雍卻璩桑桂濮牛寿通边扈燕冀郏浦尚农"
    "温别庄晏柴瞿阎充慕连茹习宦艾鱼容向古易慎戈廖庾终暨居衡步都耿满弘"
    "匡国文寇广禄阙东欧殳沃利蔚越夔隆师巩厍聂晁勾敖融冷訾辛阚那简饶空"
    "曾毋沙乜养鞠须丰巢关蒯相查后荆红游竺权逯盖益桓公"
)
COMPOUND = (
    "欧阳", "太史", "端木", "上官", "司马", "东方", "独孤", "南宫",
    "万俟", "闻人", "夏侯", "诸葛", "尉迟", "公孙", "慕容", "长孙",
    "宇文", "司徒", "轩辕", "令狐",
)
NAME_FIELDS = {"searchClientName"}
CATALOG_FIELDS = {
    "polNoInfo.plancodeinfo.abbrname",
    "pCategorys",
    "contactOrHomeAddress",
    "pajjMemberGradeInfo.pajjmemberstatus",
    "pajjMemberGradeInfo.pajjmemberproductname",
    "clientNo",
    "polNo",
}


def parse_query(raw: object) -> str:
    if isinstance(raw, str) and raw.strip().startswith("{"):
        try:
            data = json.loads(raw)
            return str(data.get("user_text") or data.get("query") or "").strip()
        except json.JSONDecodeError:
            pass
    match = re.search(r'"user_text"\s*:\s*"([^"]+)"', str(raw or ""))
    return match.group(1) if match else ""


def parse_conditions(raw: object) -> list[dict]:
    if not isinstance(raw, str) or not raw.strip().startswith("{"):
        return []
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return []
    conds = data.get("conditions")
    return [item for item in conds if isinstance(item, dict)] if isinstance(conds, list) else []


def has_surname_shape(query: str) -> bool:
    if not BARE_NAME.fullmatch(query):
        return False
    if any(query.startswith(item) for item in COMPOUND):
        return True
    return query[0] in SURNAMES


def load_set_a() -> list[dict]:
    workbook = load_workbook(NEW_XLSX, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    for raw_row in sheet.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, raw_row))
        query = parse_query(rec.get("Input / Live Request"))
        conds = parse_conditions(rec.get("Output / 被评估输出"))
        fields = [str(item.get("field") or "") for item in conds if item.get("field")]
        values = []
        for item in conds:
            value = item.get("value")
            if isinstance(value, list):
                values.extend(str(part) for part in value)
            elif value is not None:
                values.append(str(value))
        rows.append(
            {
                "id": rec.get("ID"),
                "query": query,
                "current": rec.get("状态"),
                "fields": fields,
                "values": values,
                "source": "set_a",
            }
        )
    return rows


def load_set_b() -> list[dict]:
    payload = json.loads(SET_B.read_text(encoding="utf-8"))
    rows = []
    for case in payload["cases"]:
        rows.append(
            {
                "id": case["id"],
                "query": case["query"],
                "current": None,
                "expected": case["expected_status"],
                "class": case["class"],
                "fields": [],
                "values": [],
                "source": "set_b",
            }
        )
    return rows


def build_catalog(set_a: list[dict]) -> dict[str, set[str]]:
    products: set[str] = set()
    addresses: set[str] = set()
    for row in set_a:
        for field, value in zip(row["fields"], row["values"] or []):
            token = str(value).strip()
            if not token:
                continue
            if field == "polNoInfo.plancodeinfo.abbrname":
                products.add(token)
            if field == "contactOrHomeAddress":
                addresses.add(token)
    # 金凤/金风 are the same spoken product in this table.
    if "金凤" in products:
        products.add("金风")
    return {"products": products, "addresses": addresses}


def is_parked(row: dict) -> bool:
    query = row["query"]
    if LAST_YEAR.search(query):
        return True
    if PARKED_QUERY.search(query):
        return True
    if KINSHIP.search(query) and "searchClientName" in row["fields"]:
        return True
    return False


def is_panke(query: str) -> bool:
    return bool(PANKE.search(query))


def catalog_blocks(query: str, row: dict, catalog: dict[str, set[str]]) -> bool:
    if any(field in CATALOG_FIELDS and field not in NAME_FIELDS for field in row["fields"]):
        # Live already used a non-name catalog field. Name exit abstains,
        # except when searchClientName is also present as a false name map.
        if "searchClientName" not in row["fields"]:
            return False
    for product in catalog["products"]:
        if product and product in query:
            return True
    for address in catalog["addresses"]:
        if address and query == address:
            return True
    if BUSINESS_NON_NAME.search(query):
        return True
    return False


def exit_wide(row: dict, catalog: dict[str, set[str]]) -> str | None:
    query = row["query"]
    if is_parked(row):
        return None
    if is_panke(query):
        return "not_fulfilled"
    if row["source"] == "set_b":
        if BARE_NAME.fullmatch(query):
            return "fulfilled"
        if PRODUCT_HINT.search(query):
            return "fulfilled"
        if ID_TOKEN.search(query):
            return "fulfilled"
        return None
    if catalog_blocks(query, row, catalog) and "searchClientName" in row["fields"]:
        return "not_fulfilled"
    if BARE_NAME.fullmatch(query):
        return "fulfilled"
    return None


def exit_surname(row: dict, catalog: dict[str, set[str]]) -> str | None:
    query = row["query"]
    if is_parked(row):
        return None
    if is_panke(query):
        return "not_fulfilled"
    if row["source"] == "set_b":
        if BARE_NAME.fullmatch(query):
            return "fulfilled" if has_surname_shape(query) else "not_fulfilled"
        if PRODUCT_HINT.search(query):
            return "fulfilled"
        if ID_TOKEN.search(query):
            return "fulfilled"
        return None
    if catalog_blocks(query, row, catalog) and (
        "searchClientName" in row["fields"] or BARE_NAME.fullmatch(query)
    ):
        # Directory/product/business token wins over name morphology.
        if "searchClientName" in row["fields"]:
            return "not_fulfilled"
        return None
    if BARE_NAME.fullmatch(query):
        if has_surname_shape(query):
            return "fulfilled"
        if "searchClientName" in row["fields"]:
            return "not_fulfilled"
        return None
    return None


def apply_exit(row: dict, exit_status: str | None) -> str | None:
    if exit_status is None:
        return row["current"]
    return exit_status


def summarize(rows: list[dict], key: str) -> dict:
    statuses = Counter(row[key] for row in rows if row.get(key))
    return dict(statuses)


def flips(rows: list[dict], before: str, after: str) -> list[dict]:
    out = []
    for row in rows:
        if row.get(before) and row.get(after) and row[before] != row[after]:
            out.append(
                {
                    "id": row["id"],
                    "query": row["query"],
                    "fields": row["fields"],
                    "from": row[before],
                    "to": row[after],
                }
            )
    return out


def gate_set_a(rows: list[dict], key: str) -> dict:
    probes = {
        "panke_or_circle": [row for row in rows if is_panke(row["query"])],
        "fake_name": [row for row in rows if row["query"] in {"共展", "豆芽", "见光", "傻生"}],
        "head_name": [row for row in rows if row["query"] in {"杨杰", "郑鑫", "匡西永", "王坤林", "昊轩"}],
        "catalog_product": [
            row
            for row in rows
            if row["query"] in {"金凤", "宝贝卡", "孝心", "满意", "陇佑智盛"}
        ],
    }
    report = {}
    for name, items in probes.items():
        report[name] = [
            {"id": row["id"], "query": row["query"], "current": row["current"], "sim": row[key]}
            for row in items
        ]
    return report


def gate_set_b(rows: list[dict], key: str) -> dict:
    return {
        "n": len(rows),
        "fulfilled": sum(1 for row in rows if row[key] == "fulfilled"),
        "not_fulfilled": sum(1 for row in rows if row[key] == "not_fulfilled"),
        "missing": [
            {"id": row["id"], "query": row["query"], "class": row.get("class"), "sim": row[key]}
            for row in rows
            if row[key] != "fulfilled"
        ],
    }


def main() -> int:
    set_a = load_set_a()
    set_b = load_set_b()
    catalog = build_catalog(set_a)
    for row in set_a:
        row["wide"] = apply_exit(row, exit_wide(row, catalog))
        row["surname"] = apply_exit(row, exit_surname(row, catalog))
        row["wide_exit"] = exit_wide(row, catalog)
        row["surname_exit"] = exit_surname(row, catalog)
    for row in set_b:
        row["wide"] = apply_exit(row, exit_wide(row, catalog))
        row["surname"] = apply_exit(row, exit_surname(row, catalog))
        row["wide_exit"] = exit_wide(row, catalog)
        row["surname_exit"] = exit_surname(row, catalog)

    payload = {
        "set_a": {
            "n": len(set_a),
            "current": summarize(set_a, "current"),
            "wide": summarize(set_a, "wide"),
            "surname": summarize(set_a, "surname"),
            "flips_wide": flips(set_a, "current", "wide"),
            "flips_surname": flips(set_a, "current", "surname"),
            "gates_wide": gate_set_a(set_a, "wide"),
            "gates_surname": gate_set_a(set_a, "surname"),
        },
        "set_b": {
            "note": "Set B has no live/judge traces. Scores are program exits, not LLM judge reruns.",
            "wide": gate_set_b(set_b, "wide"),
            "surname": gate_set_b(set_b, "surname"),
        },
        "catalog_projection": {
            "n_products": len(catalog["products"]),
            "products": sorted(catalog["products"]),
            "n_addresses": len(catalog["addresses"]),
        },
        "parked_untouched": [
            {"id": row["id"], "query": row["query"], "current": row["current"], "wide": row["wide"], "surname": row["surname"]}
            for row in set_a
            if is_parked(row)
        ],
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(
        json.dumps(
            {
                "wrote": str(OUT),
                "set_a_current": payload["set_a"]["current"],
                "set_a_wide": payload["set_a"]["wide"],
                "set_a_surname": payload["set_a"]["surname"],
                "n_flips_wide": len(payload["set_a"]["flips_wide"]),
                "n_flips_surname": len(payload["set_a"]["flips_surname"]),
                "set_b_wide_f": payload["set_b"]["wide"]["fulfilled"],
                "set_b_surname_f": payload["set_b"]["surname"]["fulfilled"],
                "set_b_surname_miss": payload["set_b"]["surname"]["missing"],
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
