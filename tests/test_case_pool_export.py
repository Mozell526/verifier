import json
import subprocess
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
EXPORT_JS = ROOT / "impl" / "frontend" / "case_pool_export.js"
SUMMARY_HTML = ROOT / "impl" / "frontend" / "summary.html"

TWO_TURN_SHOW = {
    "available": True,
    "mock": {
        "user_intent": "上一轮只点名保额，本轮用金额短答补槽",
        "query": "保额的保单",
        "scenario": "clarification_reply",
    },
    "overview": {
        "completion_status": "completed",
        "turn_count": 2,
        "final_output_turn": 2,
        "stop_reason": "goal_satisfied",
    },
    "turns": [
        {
            "turn_index": 1,
            "mock_message": "保额的保单",
            "status": "succeeded",
            "runtime_ms": 2475,
            "output": [
                {"path": "status", "found": True, "value": "UNSUPPORTED"},
                {"path": "message", "found": True, "value": "“保额”缺少必要的条件，请补充后重试"},
                {"path": "query", "found": True, "value": "保额的保单"},
                {"path": "filter", "found": True, "value": None},
            ],
        },
        {
            "turn_index": 2,
            "mock_message": "50万以上",
            "status": "succeeded",
            "runtime_ms": 26,
            "output": [
                {"path": "status", "found": True, "value": "SUCCESS"},
                {"path": "message", "found": True, "value": "解析成功"},
                {"path": "query", "found": True, "value": "50万以上"},
                {"path": "filter", "found": True, "value": {
                    "type": "PREDICATE",
                    "node_id": "n0",
                    "field": "sum_ins",
                    "operator": "GTE",
                    "value": 500000,
                }},
            ],
        },
    ],
}

EXPECTED_TWO_TURN = """\
Mock: 上一轮只点名保额，本轮用金额短答补槽
场景: clarification_reply
概览: completed · 2 轮 · 最终输出轮 2 · 停止 goal_satisfied

T1  succeeded  2475ms
    输入: 保额的保单
    status: UNSUPPORTED
    message: “保额”缺少必要的条件，请补充后重试
    query: 保额的保单
    filter: null

T2  succeeded  26ms
    输入: 50万以上
    status: SUCCESS
    message: 解析成功
    query: 50万以上
    filter: {"type":"PREDICATE","node_id":"n0","field":"sum_ins","operator":"GTE","value":500000}"""


def _eval_exporter(script: str, *, stdin: str = "") -> subprocess.CompletedProcess[str]:
    prelude = f"""
const fs = require('fs');
const vm = require('vm');
const context = {{ console, globalThis: {{}} }};
vm.createContext(context);
vm.runInContext(fs.readFileSync({json.dumps(str(EXPORT_JS))}, 'utf8'), context);
const CasePoolExporter = context.globalThis.CasePoolExporter;
"""
    return subprocess.run(
        ["node", "-e", prelude + script],
        input=stdin,
        capture_output=True,
        text=True,
        check=False,
    )


def test_exporter_adds_trace_summary_as_last_column():
    result = _eval_exporter("process.stdout.write(JSON.stringify(CasePoolExporter.COLUMNS));")
    assert result.returncode == 0, result.stderr
    columns = json.loads(result.stdout)
    assert columns[-1]["header"] == "Trace 摘要"
    assert columns[-1]["key"] == "traceSummary"


def test_format_trace_show_returns_no_trace_when_projection_missing():
    result = _eval_exporter(
        "process.stdout.write(CasePoolExporter.formatTraceShow(null));"
    )
    assert result.returncode == 0, result.stderr
    assert result.stdout == "无 Trace"
    result = _eval_exporter(
        "process.stdout.write(CasePoolExporter.formatTraceShow({available:false}));"
    )
    assert result.returncode == 0, result.stderr
    assert result.stdout == "无 Trace"


def test_format_trace_show_renders_approved_two_turn_summary():
    result = _eval_exporter(
        "const payload=JSON.parse(fs.readFileSync(0,'utf8'));"
        "process.stdout.write(CasePoolExporter.formatTraceShow(payload));",
        stdin=json.dumps(TWO_TURN_SHOW, ensure_ascii=False),
    )
    assert result.returncode == 0, result.stderr
    assert result.stdout == EXPECTED_TWO_TURN


def test_summary_export_rows_pull_trace_show_into_trace_summary():
    source = SUMMARY_HTML.read_text(encoding="utf-8")
    builder = source[source.index("function buildCasePoolExportRows"):source.index("async function exportCurrentCasePoolXlsx")]
    formatter = source[source.index("function exportTraceSummary"):source.index("function buildCasePoolExportRows")]
    assert "traceSummary:exportTraceSummary(view)" in builder.replace(" ", "")
    assert "item.frontend_view?.project_extensions?.trace_show" in formatter
    assert "CasePoolExporter.formatTraceShow" in formatter
    assert "typeof formatter!=='function'" in formatter
    assert "item.trace" not in formatter
    assert "case_pool_export.js?v=20260813-trace-summary-1" in source


def test_exporter_writes_xlsx_with_trace_summary_column(tmp_path):
    out = tmp_path / "trace-export.xlsx"
    result = subprocess.run(
        ["node", str(ROOT / "tests" / "write_trace_export_xlsx.js"), str(out)],
        capture_output=True,
        text=True,
        check=False,
    )
    assert result.returncode == 0, result.stderr
    assert out.exists() and out.stat().st_size > 0
    from openpyxl import load_workbook
    workbook = load_workbook(out, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert headers[-1] == "Trace 摘要"
    values = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    summary = values[-1]
    assert "T1  succeeded  2475ms" in summary
    assert "T2  succeeded  26ms" in summary
    assert "输入: 50万以上" in summary
    assert "status: SUCCESS" in summary
    workbook.close()
