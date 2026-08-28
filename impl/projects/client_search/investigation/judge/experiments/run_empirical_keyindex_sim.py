#!/usr/bin/env python3
"""Empirical offline simulation of field/multi-collection key-index variants.

V0: current field Builder projection + char-heuristic
V1: field projection + examples/notes (source-derived) + fused exact/lexical
V2: V1 field + enhanced_rules pattern index + abbrname exact membership
V3: V1 + 2-char stem / contained-phrase when longer phrase in source projection

Evaluation-only. No Draft skill edits. No selected/promote/solidify claims.
"""
from __future__ import annotations

import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field as dc_field
from pathlib import Path
from typing import Any, Callable, Sequence

import yaml

ROOT = Path("/workspace/draft-run")
SRC = ROOT / "source"
XLSX = Path("/workspace/verifier-client_search-cases-20260812-214445.xlsx")
OUT_JSON = ROOT / "empirical-keyindex-sim.json"
OUT_MD = ROOT / "empirical-keyindex-sim.md"
ISSUES = ROOT / "issues-log.md"

_IGNORED_QUERY_CHARS = set("客户的有是和与及或并且一个哪些名单帮我找查询")


@dataclass
class Entry:
    key: str
    name: str
    search_text: str
    target_ref: str
    collection: str
    payload: dict[str, Any] = dc_field(default_factory=dict)


def _normalise(value: Any) -> str:
    text = "".join(
        char
        for char in str(value or "").casefold()
        if char.isalnum() or "\u4e00" <= char <= "\u9fff"
    )
    return re.sub(r"\d+", "0", text)


def _bigrams(value: Any) -> set[str]:
    text = _normalise(value)
    return {text[i : i + 2] for i in range(max(0, len(text) - 1))}


def _phrases(value: Any) -> list[str]:
    out = []
    for part in re.split(r"[\s，。；、：:（）()|/]+", str(value or "")):
        text = _normalise(part)
        if len(text) >= 2:
            out.append(text)
    return out


def _short_name(value: Any) -> str:
    short_name = str(value or "").strip().split("，", 1)[0].split("。", 1)[0].strip()
    for prefix in ("仅表示", "表示"):
        if short_name.startswith(prefix):
            short_name = short_name[len(prefix) :].strip()
            break
    return short_name[:32]


def _searchable_chars(value: Any) -> set[str]:
    return {
        char
        for char in str(value or "").casefold()
        if (
            not char.isspace()
            and char not in _IGNORED_QUERY_CHARS
            and (char.isalpha() or "\u4e00" <= char <= "\u9fff")
        )
    }


def _example_queries(raw: dict[str, Any]) -> list[str]:
    queries = []
    for ex in raw.get("examples") or []:
        if isinstance(ex, dict):
            q = ex.get("query")
            if q:
                queries.append(str(q))
        elif isinstance(ex, str) and ex.strip():
            queries.append(ex.strip())
    return queries


def build_field_entries_v0(intents: list[dict[str, Any]]) -> list[Entry]:
    entries: list[Entry] = []
    seen: set[str] = set()
    for raw in intents:
        if not isinstance(raw, dict):
            continue
        field = str(raw.get("field") or "").strip()
        if not field or field in seen:
            continue
        seen.add(field)
        search_text = " ".join(
            str(raw.get(k) or "") for k in ("field", "retrieval_text", "description")
        )
        entries.append(
            Entry(
                key=field,
                name=_short_name(raw.get("description")) or field,
                search_text=search_text,
                target_ref=f"field://{field}",
                collection="field",
                payload={"field": field, "projection": "v0"},
            )
        )
    return entries


def build_field_entries_v1(intents: list[dict[str, Any]]) -> list[Entry]:
    """Aggregate ALL intents per field; project examples/notes (source-derived)."""
    buckets: dict[str, dict[str, Any]] = {}
    for raw in intents:
        if not isinstance(raw, dict):
            continue
        field = str(raw.get("field") or "").strip()
        if not field:
            continue
        b = buckets.setdefault(
            field,
            {
                "field": field,
                "description": "",
                "parts": [],
                "examples": [],
                "notes": [],
                "retrieval_texts": [],
            },
        )
        if not b["description"] and raw.get("description"):
            b["description"] = str(raw.get("description"))
        if raw.get("retrieval_text"):
            b["retrieval_texts"].append(str(raw["retrieval_text"]))
        if raw.get("notes"):
            b["notes"].append(str(raw["notes"]))
        b["examples"].extend(_example_queries(raw))
        b["parts"].append(str(raw.get("field") or ""))
    entries = []
    for field, b in buckets.items():
        # de-dupe preserving order
        def uniq(xs):
            seen = set()
            out = []
            for x in xs:
                x = str(x).strip()
                if not x or x in seen:
                    continue
                seen.add(x)
                out.append(x)
            return out

        search_text = " ".join(
            uniq(
                [field]
                + b["retrieval_texts"]
                + ([b["description"]] if b["description"] else [])
                + b["notes"]
                + b["examples"]
            )
        )
        entries.append(
            Entry(
                key=field,
                name=_short_name(b["description"]) or field,
                search_text=search_text,
                target_ref=f"field://{field}",
                collection="field",
                payload={
                    "field": field,
                    "projection": "v1",
                    "n_examples": len(uniq(b["examples"])),
                    "n_notes": len(uniq(b["notes"])),
                },
            )
        )
    return entries


def build_rule_entries(rules: list[dict[str, Any]]) -> list[Entry]:
    entries = []
    for i, rule in enumerate(rules):
        if not isinstance(rule, dict):
            continue
        name = str(rule.get("name") or f"rule-{i}").strip()
        patterns = rule.get("patterns") or []
        if isinstance(patterns, str):
            patterns = [patterns]
        pat_text = " ".join(str(p) for p in patterns)
        field = str(rule.get("field") or "").strip()
        search_text = " ".join(x for x in (name, pat_text, field) if x)
        # compact load payload: rule subtree, not full collection
        compact = {
            k: rule.get(k)
            for k in (
                "name",
                "patterns",
                "field",
                "operator",
                "match_mode",
                "value_type",
                "value",
                "priority",
                "query_logic",
                "merge_to_llm",
                "extra_conditions",
            )
            if k in rule
        }
        entries.append(
            Entry(
                key=name,
                name=name,
                search_text=search_text,
                target_ref=f"rule://{name}",
                collection="enhanced_rules",
                payload=compact,
            )
        )
    return entries


def build_abbrname_exact(doc: dict[str, Any]) -> set[str]:
    vals = []
    node = doc.get("polNoInfo.plancodeinfo.abbrname") or doc
    if isinstance(node, dict):
        vals = node.get("values") or node.get("enums") or []
    elif isinstance(node, list):
        vals = node
    return {_normalise(v) for v in vals if str(v).strip()}


# --- strategies ---

def char_heuristic(query: str, entries: Sequence[Entry], limit: int):
    query_text = str(query or "").strip()
    query_chars = _searchable_chars(query_text)
    if not query_chars:
        return []
    candidates = []
    for entry in entries:
        score = len(query_chars & _searchable_chars(entry.search_text))
        if entry.key.casefold() in query_text.casefold():
            score += 100
        if "岁" in query_text or "周岁" in query_text:
            if "年龄" in entry.name or "age" in entry.key.casefold():
                score += 20
            if "family" in entry.key.casefold():
                score -= 5
        if score >= 2:
            candidates.append((entry, float(score), {"lexical": float(score)}))
    candidates.sort(key=lambda item: (-item[1], item[0].key))
    return candidates[:limit]


def exact_strategy(query: str, entries: Sequence[Entry], limit: int, *, allow_2char_stem: bool = False):
    raw_query = str(query or "").casefold()
    query_text = _normalise(query)
    ranked = []
    for entry in entries:
        scores = []
        if entry.key.casefold() in raw_query:
            scores.append(100.0)
        for phrase in _phrases(entry.search_text):
            if phrase == query_text or phrase in query_text:
                scores.append(float(20 + min(len(phrase), 20)))
            elif allow_2char_stem and len(query_text) >= 2 and query_text in phrase and len(phrase) > len(query_text):
                # only when longer phrase literally in source projection
                scores.append(float(15 + min(len(query_text), 10)))
            elif (not allow_2char_stem) and len(query_text) >= 3 and query_text in phrase:
                scores.append(float(15 + min(len(query_text), 10)))
        if scores:
            ranked.append((entry, max(scores), {"exact": max(scores)}))
    ranked.sort(key=lambda item: (-item[1], item[0].key))
    return ranked[:limit]


def source_phrase_idf_strategy(
    entries: Sequence[Entry],
    *,
    min_query_coverage: float = 0.0,
    allow_2char_stem: bool = False,
):
    document_frequency: Counter[str] = Counter()
    for entry in entries:
        document_frequency.update(_bigrams(entry.search_text))
    count = len(entries)

    def idf(term: str) -> float:
        return math.log((count + 1) / (document_frequency[term] + 1)) + 1

    def search(query: str, candidates: Sequence[Entry], limit: int):
        query_text = _normalise(query)
        query_bigrams = _bigrams(query_text)
        ranked = []
        if len(query_text) < 2:
            return ranked
        for entry in candidates:
            exact_phrases = []
            for phrase in _phrases(entry.search_text):
                if phrase == query_text or (len(phrase) >= 3 and phrase in query_text):
                    exact_phrases.append(phrase)
                elif allow_2char_stem and len(query_text) >= 2 and query_text in phrase and len(phrase) > len(query_text):
                    exact_phrases.append(phrase)
                elif (not allow_2char_stem) and len(query_text) >= 3 and query_text in phrase:
                    exact_phrases.append(phrase)
            shared = query_bigrams & _bigrams(entry.search_text)
            key_hit = entry.key.casefold() in str(query or "").casefold()
            coverage = len(shared) / len(query_bigrams) if query_bigrams else 0.0
            if not key_hit and not exact_phrases and (len(shared) < 2 or coverage < min_query_coverage):
                continue
            score = (
                sum(idf(term) for term in sorted(shared))
                + sum(8 + min(len(phrase), 12) for phrase in exact_phrases)
                + (100 * coverage)
                + (100 if key_hit else 0)
            )
            ranked.append((entry, float(score), {"lexical": float(score)}))
        ranked.sort(key=lambda item: (-item[1], item[0].key))
        return ranked[:limit]

    return search


def fused_strategy(exact_fn, lexical_fn):
    def search(query: str, entries: Sequence[Entry], limit: int):
        merged: dict[str, tuple[Entry, dict[str, float]]] = {}
        for strategy in (exact_fn, lexical_fn):
            for entry, _score, channels in strategy(query, entries, max(limit, 8)):
                existing = merged.get(entry.key)
                combined = dict(existing[1]) if existing else {}
                combined.update(channels)
                merged[entry.key] = (entry, combined)
        ranked = []
        for entry, channels in merged.values():
            score = (1000 if "exact" in channels else 0) + sum(channels[k] for k in sorted(channels))
            ranked.append((entry, float(score), channels))
        ranked.sort(key=lambda item: (-item[1], item[0].key))
        return ranked[:limit]

    return search


def rules_search(query: str, entries: Sequence[Entry], limit: int):
    """Exact/near-exact on rule name + patterns.

    Directionality matters:
    - allow query⊂pattern (有钱 in 有钱客户/高价值 patterns)
    - do NOT allow short generic pattern⊂query (客户 in 关爱客户) unless pattern==query
      or pattern length >= 4.
    """
    qn = _normalise(query)
    if len(qn) < 1:
        return []
    ranked = []
    for entry in entries:
        scores = []
        name_n = _normalise(entry.key)
        pats = entry.payload.get("patterns") or []
        if isinstance(pats, str):
            pats = [pats]
        pat_norms = [_normalise(p) for p in pats if str(p).strip()]
        if name_n == qn:
            scores.append(80.0)
        elif len(qn) >= 4 and qn in name_n:
            scores.append(55.0)
        for pn in pat_norms:
            if not pn:
                continue
            if pn == qn:
                scores.append(70.0)
            elif len(qn) >= 2 and qn in pn:
                # query contained in a source pattern (stem OK)
                scores.append(float(45 + min(len(qn), 20)))
            elif len(pn) >= 4 and pn in qn:
                # longer pattern appears inside the user query
                scores.append(float(35 + min(len(pn), 20)))
        if re.fullmatch(r"[a-z0-9]+", qn) and len(qn) <= 4:
            ok = any(pn == qn or (len(qn) >= 2 and qn in pn) for pn in pat_norms)
            if not ok:
                for p in pats:
                    for tok in re.split(r"[\s|/，,、]+", str(p)):
                        if _normalise(tok) == qn:
                            ok = True
                            break
            if not ok:
                continue
            if not scores:
                scores.append(60.0)
        if not scores:
            continue
        ranked.append((entry, max(scores), {"exact": max(scores)}))
    ranked.sort(key=lambda item: (-item[1], item[0].key))
    return ranked[:limit]


def abbrname_exact_lookup(query: str, members: set[str], limit: int = 8):
    qn = _normalise(query)
    if not qn or qn not in members:
        return []
    entry = Entry(
        key=qn,
        name=qn,
        search_text=qn,
        target_ref=f"abbrname://{qn}",
        collection="abbrname",
        payload={"value": qn, "field": "polNoInfo.plancodeinfo.abbrname", "membership": "exact"},
    )
    return [(entry, 100.0, {"exact": 100.0})]


# --- evaluation helpers ---

FOCUS = [
    ("focus-guanai", "关爱客户", "business_noun", {"fields": {"clientAge"}, "rules_substr": ["关爱客户"], "abbr": set()}),
    ("focus-youqian", "有钱", "business_noun", {"fields": {"newValueLabel"}, "rules_substr": ["有钱", "高价值"], "abbr": set()}),
    ("focus-youqian-kehu", "有钱客户", "business_noun", {"fields": {"newValueLabel"}, "rules_substr": ["有钱", "高价值"], "abbr": set()}),
    ("focus-jinfeng", "金凤", "business_noun", {"fields": set(), "rules_substr": [], "abbr": {"金凤", "金凤100"}}),
    ("focus-panke", "盘客", "business_noun", {"fields": {"customerReview"}, "rules_substr": ["盘客"], "abbr": set()}),
    ("focus-qupanke", "去盘客", "business_noun", {"fields": {"customerReview"}, "rules_substr": ["盘客"], "abbr": set()}),
    ("focus-latin-a", "A", "latin_bareword", {"fields": set(), "rules_substr": [], "abbr": set(), "reject_preferred": True}),
    ("focus-latin-o2o", "O2O", "latin_bareword", {"fields": {"pcustSourcType", "validSinsPol"}, "rules_substr": ["O2O", "准客来源"], "abbr": set()}),
    ("focus-name-chen", "陈金秀", "person_name", {"fields": {"searchClientName", "polNoInfo.applicantname", "polNoInfo.insuredname"}, "rules_substr": [], "abbr": set(), "name_ok_fields": True}),
    ("focus-weather", "天气怎么样", "irrelevant_like", {"fields": set(), "rules_substr": [], "abbr": set(), "reject_preferred": True}),
    ("focus-hobby", "客户平时有什么兴趣爱好", "unsupported_like", {"fields": set(), "rules_substr": [], "abbr": set(), "reject_preferred": True}),
]


def _bucketize_query(q: str) -> str:
    n = _normalise(q)
    if re.fullmatch(r"[a-z0-9]+", n) and len(n) <= 4:
        return "latin_bareword"
    if any(x in q for x in ("天气", "红烧肉", "黑洞", "咖啡", "银河", "写一首诗", "量子")):
        return "irrelevant_like"
    if any(x in q for x in ("兴趣爱好", "喜欢的颜色", "宠物", "电影")):
        return "unsupported_like"
    if any(x in q for x in ("关爱客户", "有钱", "金凤", "盘客", "高价值", "续收")):
        return "business_noun"
    if re.search(r"[\u4e00-\u9fff]{2,3}", q) and any(k in q for k in ("姓名", "叫", "陈", "客户是")):
        return "person_name_context"
    return "other"


def judge_focus(query: str, meta: dict, hits: list[dict[str, Any]]) -> str:
    if meta.get("reject_preferred"):
        return "OK" if not hits else "FALSE"
    fields = {h["key"] for h in hits if h["collection"] == "field"}
    rules = {h["key"] for h in hits if h["collection"] == "enhanced_rules"}
    abbr = {h["key"] for h in hits if h["collection"] == "abbrname"}
    want_fields = set(meta.get("fields") or [])
    if want_fields & fields:
        return "OK"
    for sub in meta.get("rules_substr") or []:
        if any(sub in r for r in rules):
            return "OK"
    want_abbr = {_normalise(x) for x in (meta.get("abbr") or set())}
    if want_abbr & abbr:
        return "OK"
    # person name: empty is OK-ish (MISS for navigation); hitting unrelated = FALSE; name fields = OK
    if meta.get("name_ok_fields"):
        if want_fields & fields:
            return "OK"
        if not hits:
            return "MISS"
        return "FALSE"
    if not hits:
        return "MISS"
    # hit something but not relevant target
    return "FALSE"


def load_xlsx_queries(path: Path) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError:
        return []
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h or "").strip() for h in next(rows)]

    def col(*names):
        lower = {h.lower(): i for i, h in enumerate(header)}
        for n in names:
            if n.lower() in lower:
                return lower[n.lower()]
        for i, h in enumerate(header):
            for n in names:
                if n.lower() in h.lower():
                    return i
        return None

    iq = col("user_text", "query", "input", "用户输入", "query_text")
    ifields = col("output_fields", "fields", "actual_fields")
    out = []
    for i, row in enumerate(rows):
        if iq is None or iq >= len(row):
            continue
        q = row[iq]
        if q is None or not str(q).strip():
            continue
        item = {"id": f"xlsx-{i+1}", "query": str(q).strip(), "bucket": "xlsx_stress", "required": []}
        if ifields is not None and ifields < len(row) and row[ifields]:
            raw = row[ifields]
            if isinstance(raw, str):
                item["weak_live_fields"] = [p.strip() for p in re.split(r"[,;|\s]+", raw) if p.strip()]
        out.append(item)
    return out


def load_iteration_queries(path: Path) -> list[dict[str, Any]]:
    cases = json.loads(path.read_text(encoding="utf-8"))
    out = []
    for case in cases:
        cid = case.get("id") or case.get("source_case_id")
        trace = case.get("trace") or {}
        extracted = trace.get("extracted_output") or {}
        q = extracted.get("query") or (trace.get("input") or {}).get("query")
        if not q:
            inp = trace.get("input") or {}
            q = inp.get("user_text") or inp.get("text")
        if not q:
            continue
        fields = []
        for cond in extracted.get("conditions") or []:
            if isinstance(cond, dict) and cond.get("field"):
                fields.append(str(cond["field"]))
        out.append(
            {
                "id": f"iter-{cid}",
                "query": str(q),
                "bucket": "iteration_case",
                "required": [],
                "weak_live_fields": fields,
            }
        )
    return out


def run_search(variant: str, query: str, ctx: dict[str, Any], limit: int = 8) -> dict[str, Any]:
    hits: list[tuple[Entry, float, dict[str, float]]] = []
    if variant == "V0":
        hits = char_heuristic(query, ctx["field_v0"], limit)
    elif variant == "V1":
        hits = ctx["field_v1_fused"](query, ctx["field_v1"], limit)
    elif variant == "V3":
        hits = ctx["field_v3_fused"](query, ctx["field_v1"], limit)
    elif variant == "V2":
        field_hits = ctx["field_v1_fused"](query, ctx["field_v1"], limit)
        rule_hits = rules_search(query, ctx["rules"], limit)
        abbr_hits = abbrname_exact_lookup(query, ctx["abbr_members"], limit)
        # Slot merge: prefer exact abbr/rules, keep field slots so labeled field recall
        # is not drowned by over-broad rule lists. Still Search→load objects, not dumps.
        hits = []
        seen = set()
        for batch, cap in (
            (abbr_hits, 2),
            (rule_hits, 3),
            (field_hits, limit),
        ):
            taken = 0
            for entry, score, channels in batch:
                mk = f"{entry.collection}:{entry.key}"
                if mk in seen:
                    continue
                hits.append((entry, float(score), channels))
                seen.add(mk)
                taken += 1
                if taken >= cap or len(hits) >= limit:
                    break
            if len(hits) >= limit:
                break
        hits = hits[:limit]
    else:
        raise ValueError(variant)

    loaded = []
    for entry, score, channels in hits:
        obj = None
        if entry.collection == "field":
            obj = ctx["field_defs"].get(entry.key) or {"field": entry.key}
        elif entry.collection == "enhanced_rules":
            obj = entry.payload
        elif entry.collection == "abbrname":
            obj = entry.payload
        loaded.append(
            {
                "key": entry.key,
                "collection": entry.collection,
                "score": score,
                "channels": channels,
                "target_ref": entry.target_ref,
                "loaded": obj is not None,
                "chars": len(json.dumps(obj, ensure_ascii=False)) if obj is not None else 0,
            }
        )
    return {
        "hit_keys": [x["key"] for x in loaded],
        "hit_collections": [x["collection"] for x in loaded],
        "hits": loaded,
        "hit_count": len(loaded),
        "loaded_count": sum(1 for x in loaded if x["loaded"]),
        "load_errors": [],
        "average_loaded_chars": (sum(x["chars"] for x in loaded) / len(loaded)) if loaded else 0.0,
    }


def metrics_labeled(rows: list[dict[str, Any]]) -> dict[str, Any]:
    relevant = [r for r in rows if r.get("required")]
    irr = [r for r in rows if r.get("category") in {"irrelevant", "unsupported"}]
    top8_hits = 0
    for r in relevant:
        req = set(r["required"])
        # for multi-collection, field keys in hits
        hits = set(r["result"]["hit_keys"])
        if req and req.issubset(hits):
            top8_hits += 1
    irr_reject = sum(1 for r in irr if r["result"]["hit_count"] == 0)
    s2l_ok = 0
    s2l_n = 0
    for r in rows:
        if r["result"]["hit_count"] == 0:
            continue
        s2l_n += 1
        if r["result"]["loaded_count"] == r["result"]["hit_count"]:
            s2l_ok += 1
    return {
        "relevant_n": len(relevant),
        "top8_recall_rate": (top8_hits / len(relevant)) if relevant else None,
        "irrelevant_n": len(irr),
        "irrelevant_rejection_rate": (irr_reject / len(irr)) if irr else None,
        "search_to_load_resolution_rate": (s2l_ok / s2l_n) if s2l_n else None,
        "queries_with_hits": sum(1 for r in rows if r["result"]["hit_count"] > 0),
        "queries_total": len(rows),
        "empty_hit_rate": sum(1 for r in rows if r["result"]["hit_count"] == 0) / len(rows) if rows else None,
        "avg_hit_count": sum(r["result"]["hit_count"] for r in rows) / len(rows) if rows else None,
    }


def main() -> int:
    field_doc = yaml.safe_load((SRC / "field_definitions_args.yaml").read_text(encoding="utf-8")) or {}
    intents = [x for x in (field_doc.get("intents") or []) if isinstance(x, dict)]
    rules_doc = yaml.safe_load((SRC / "enhanced_rules_args.yaml").read_text(encoding="utf-8")) or {}
    rules = [x for x in (rules_doc.get("rules") or []) if isinstance(x, dict)]
    abbr_doc = yaml.safe_load((SRC / "abbrname_enums_args.yaml").read_text(encoding="utf-8")) or {}

    field_v0 = build_field_entries_v0(intents)
    field_v1 = build_field_entries_v1(intents)
    rule_entries = build_rule_entries(rules)
    abbr_members = build_abbrname_exact(abbr_doc)

    # field def load map: first intent per field (full object)
    field_defs: dict[str, Any] = {}
    for raw in intents:
        f = str(raw.get("field") or "").strip()
        if f and f not in field_defs:
            field_defs[f] = raw

    idf_v1 = source_phrase_idf_strategy(field_v1, min_query_coverage=0.0, allow_2char_stem=False)
    idf_v3 = source_phrase_idf_strategy(field_v1, min_query_coverage=0.0, allow_2char_stem=True)
    exact_v1 = lambda q, e, lim: exact_strategy(q, e, lim, allow_2char_stem=False)
    exact_v3 = lambda q, e, lim: exact_strategy(q, e, lim, allow_2char_stem=True)

    ctx = {
        "field_v0": field_v0,
        "field_v1": field_v1,
        "rules": rule_entries,
        "abbr_members": abbr_members,
        "field_defs": field_defs,
        "field_v1_fused": fused_strategy(exact_v1, idf_v1),
        "field_v3_fused": fused_strategy(exact_v3, idf_v3),
    }

    # projection smoke checks
    smoke = {}
    for label, entries in (("v0", field_v0), ("v1", field_v1)):
        by = {e.key: e for e in entries}
        smoke[label] = {
            "n_entries": len(entries),
            "关爱客户_in_clientAge": "关爱客户" in (by.get("clientAge").search_text if by.get("clientAge") else ""),
            "有钱客户_in_newValueLabel": "有钱客户" in (by.get("newValueLabel").search_text if by.get("newValueLabel") else ""),
            "盘客_in_customerReview": "盘客" in (by.get("customerReview").search_text if by.get("customerReview") else ""),
            "金凤_anywhere": any("金凤" in e.search_text for e in entries),
        }
    smoke["rules"] = {
        "n": len(rule_entries),
        "关爱客户": any("关爱客户" in e.search_text for e in rule_entries),
        "有钱": any("有钱" in e.search_text for e in rule_entries),
        "盘客": any("盘客" in e.search_text for e in rule_entries),
        "O2O": any("O2O" in e.search_text or "o2o" in e.search_text.casefold() for e in rule_entries),
    }
    smoke["abbr"] = {"n": len(abbr_members), "金凤": _normalise("金凤") in abbr_members}

    dev = json.loads((SRC / "development_probes.json").read_text(encoding="utf-8"))
    hold = json.loads((SRC / "holdout_probes.json").read_text(encoding="utf-8"))
    labeled = [{**p, "split": "development"} for p in dev] + [{**p, "split": "holdout"} for p in hold]

    # I078/I036/I210-style extras from iteration
    iter_q = load_iteration_queries(SRC / "iteration-cases.json")
    style_extra = []
    for item in iter_q:
        cid = item["id"]
        if any(x in cid for x in ("078", "036", "210")):
            style_extra.append(item)

    focus_items = []
    for fid, q, bucket, meta in FOCUS:
        focus_items.append({"id": fid, "query": q, "bucket": bucket, "meta": meta, "required": []})
    for item in style_extra:
        focus_items.append(
            {
                "id": f"focus-style-{item['id']}",
                "query": item["query"],
                "bucket": "iteration_style",
                "meta": {"fields": set(item.get("weak_live_fields") or []), "rules_substr": [], "abbr": set()},
                "required": [],
                "weak_live_fields": item.get("weak_live_fields") or [],
            }
        )

    stress = list(iter_q)
    stress.extend(load_xlsx_queries(XLSX)[:341])
    for fid, q, bucket, meta in FOCUS:
        stress.append({"id": fid, "query": q, "bucket": bucket, "required": []})
    seen = set()
    deduped = []
    for item in stress:
        if item["query"] in seen:
            continue
        seen.add(item["query"])
        deduped.append(item)
    stress = deduped

    variants = ["V0", "V1", "V2", "V3"]
    report: dict[str, Any] = {
        "schema_version": 1,
        "experiment_id": "empirical-keyindex-sim-20260813",
        "note": "Offline empirical sim only. Not selected/promoted/solidified. Source-derived projections; no AI synonyms.",
        "smoke": smoke,
        "entry_counts": {
            "field_v0": len(field_v0),
            "field_v1": len(field_v1),
            "rules": len(rule_entries),
            "abbrname_members": len(abbr_members),
        },
        "query_counts": {
            "labeled_frozen": len(labeled),
            "stress": len(stress),
            "focus": len(FOCUS),
            "style_extra": len(style_extra),
        },
        "variants": {},
    }

    for variant in variants:
        labeled_rows = []
        for probe in labeled:
            result = run_search(variant, probe["query"], ctx)
            req = set(probe.get("required") or [])
            hits = set(result["hit_keys"])
            labeled_rows.append(
                {
                    "id": probe["id"],
                    "split": probe["split"],
                    "category": probe["category"],
                    "query": probe["query"],
                    "required": list(req),
                    "pass_top8": bool(req) and req.issubset(hits),
                    "rejected_ok": (not req) and result["hit_count"] == 0,
                    "result": {
                        "hit_keys": result["hit_keys"],
                        "hit_collections": result["hit_collections"],
                        "hit_count": result["hit_count"],
                        "loaded_count": result["loaded_count"],
                        "load_errors": result["load_errors"],
                        "average_loaded_chars": result["average_loaded_chars"],
                    },
                }
            )

        focus_rows = []
        for item in focus_items:
            result = run_search(variant, item["query"], ctx)
            meta = item.get("meta") or {}
            # normalize meta sets for judge
            meta_norm = {
                "fields": set(meta.get("fields") or []),
                "rules_substr": list(meta.get("rules_substr") or []),
                "abbr": set(meta.get("abbr") or []),
                "reject_preferred": bool(meta.get("reject_preferred")),
                "name_ok_fields": bool(meta.get("name_ok_fields")),
            }
            label = judge_focus(item["query"], meta_norm, result["hits"])
            focus_rows.append(
                {
                    "id": item["id"],
                    "query": item["query"],
                    "bucket": item["bucket"],
                    "qual": label,
                    "hits": [
                        {"key": h["key"], "collection": h["collection"], "score": h["score"]}
                        for h in result["hits"][:8]
                    ],
                    "loaded_count": result["loaded_count"],
                    "hit_count": result["hit_count"],
                }
            )

        stress_rows = []
        bucket_stats: dict[str, Counter] = defaultdict(Counter)
        false_accept_proxy = Counter()
        for item in stress:
            result = run_search(variant, item["query"], ctx)
            bucket = item.get("bucket") or _bucketize_query(item["query"])
            if bucket == "xlsx_stress":
                bucket = _bucketize_query(item["query"])
            weak = set(item.get("weak_live_fields") or [])
            weak_hit = len(weak & set(result["hit_keys"])) if weak else None
            stress_rows.append(
                {
                    "id": item["id"],
                    "bucket": bucket,
                    "query": item["query"],
                    "weak_live_hit_count": weak_hit,
                    "result": {
                        "hit_keys": result["hit_keys"][:8],
                        "hit_collections": result["hit_collections"][:8],
                        "hit_count": result["hit_count"],
                        "loaded_count": result["loaded_count"],
                    },
                }
            )
            bucket_stats[bucket]["n"] += 1
            bucket_stats[bucket]["with_hits"] += int(result["hit_count"] > 0)
            bucket_stats[bucket]["empty"] += int(result["hit_count"] == 0)
            if weak:
                bucket_stats[bucket]["weak_labeled"] += 1
                bucket_stats[bucket]["weak_any_hit"] += int((weak_hit or 0) > 0)
            # light false-accept proxies
            if bucket in {"irrelevant_like", "unsupported_like"} and result["hit_count"] > 0:
                false_accept_proxy[bucket] += 1

        by_split = {}
        for split in ("development", "holdout"):
            rows = [r for r in labeled_rows if r["split"] == split]
            by_split[split] = metrics_labeled(rows)
            # field-channel-only view (ignore rule/abbr keys) for multi-collection fairness
            field_only_rows = []
            for r in rows:
                rr = dict(r)
                res = dict(r["result"])
                pairs = list(zip(res.get("hit_keys") or [], res.get("hit_collections") or []))
                fk = [k for k, c in pairs if c == "field"]
                res["hit_keys"] = fk
                res["hit_count"] = len(fk)
                res["loaded_count"] = len(fk)
                rr["result"] = res
                field_only_rows.append(rr)
            by_split[split + "_field_only"] = metrics_labeled(field_only_rows)

        # solves old?
        focus_map = {r["query"]: r for r in focus_rows if r["id"].startswith("focus-") and not r["id"].startswith("focus-style")}
        flips = {
            q: focus_map[q]["qual"]
            for q in ("关爱客户", "有钱", "金凤", "盘客")
            if q in focus_map
        }

        report["variants"][variant] = {
            "labeled_metrics": by_split,
            "labeled_failures": [
                {
                    "id": r["id"],
                    "split": r["split"],
                    "category": r["category"],
                    "query": r["query"],
                    "required": r["required"],
                    "hits": r["result"]["hit_keys"],
                    "collections": r["result"]["hit_collections"],
                }
                for r in labeled_rows
                if (r["required"] and not r["pass_top8"]) or ((not r["required"]) and not r["rejected_ok"])
            ],
            "focus": focus_rows,
            "focus_flips": flips,
            "stress_summary": {
                "n": len(stress_rows),
                "with_hits": sum(1 for r in stress_rows if r["result"]["hit_count"] > 0),
                "empty_rate": sum(1 for r in stress_rows if r["result"]["hit_count"] == 0) / len(stress_rows),
                "avg_hits": sum(r["result"]["hit_count"] for r in stress_rows) / len(stress_rows),
                "load_success_rate": (
                    sum(
                        1
                        for r in stress_rows
                        if r["result"]["hit_count"]
                        and r["result"]["loaded_count"] == r["result"]["hit_count"]
                    )
                    / max(1, sum(1 for r in stress_rows if r["result"]["hit_count"] > 0))
                ),
                "false_accept_proxy": dict(false_accept_proxy),
                "buckets": {k: dict(v) for k, v in sorted(bucket_stats.items())},
            },
            # keep compact stress: only empties/false-ish samples? store all ids heavy — store summary + focus only; full stress hit_counts in buckets
            "stress_sample_false_accepts": [
                {"id": r["id"], "query": r["query"], "bucket": r["bucket"], "hits": r["result"]["hit_keys"][:5]}
                for r in stress_rows
                if r["bucket"] in {"irrelevant_like", "unsupported_like"} and r["result"]["hit_count"] > 0
            ][:20],
        }

    OUT_JSON.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    # markdown
    lines = [
        "# Empirical Key-Index Simulation (V0–V3)",
        "",
        "Generated: 2026-08-13 00:20 UTC+8",
        "",
        "## Scope",
        "",
        "- Offline deterministic Search→Load style harness only.",
        "- **Not** selected / promoted / solidified. No Draft skill edits.",
        "- Source-derived projections only (field YAML examples/notes/retrieval_text; rule name/patterns/field; abbrname exact values).",
        "",
        "### Variants",
        "",
        "| ID | Definition |",
        "|---|---|",
        "| V0 | Current field Builder projection (`field+retrieval_text+description`, first intent/field) + char-heuristic |",
        "| V1 | Field projection aggregates all intents; includes `examples.query` + `notes` + retrieval_text; fused exact+IDF lexical |",
        "| V2 | V1 field index + enhanced_rules (name/patterns/field) Search→load rule subtree + abbrname **exact** membership |",
        "| V3 | V1 projection + allow 2-char stem / query⊂phrase when longer phrase literally in source projection |",
        "",
        f"- Entry counts: field_v0={len(field_v0)}, field_v1={len(field_v1)}, rules={len(rule_entries)}, abbrname={len(abbr_members)}",
        f"- Queries: labeled={len(labeled)}, stress_deduped={len(stress)}, focus={len(FOCUS)}, style_extra={len(style_extra)}",
        "",
        "### Projection smoke",
        "",
        "```json",
        json.dumps(smoke, ensure_ascii=False, indent=2),
        "```",
        "",
        "## Headline labeled metrics",
        "",
        "| variant | dev top8 | dev irr_rej | holdout top8 | holdout irr_rej | stress empty | avg hits | load_success |",
        "|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for v in variants:
        p = report["variants"][v]
        d = p["labeled_metrics"]["development"]
        h = p["labeled_metrics"]["holdout"]
        s = p["stress_summary"]
        lines.append(
            f"| {v} | {d.get('top8_recall_rate'):.3f} | {d.get('irrelevant_rejection_rate'):.3f} | "
            f"{h.get('top8_recall_rate'):.3f} | {h.get('irrelevant_rejection_rate'):.3f} | "
            f"{s['empty_rate']:.3f} | {s['avg_hits']:.2f} | {s['load_success_rate']:.3f} |"
        )

    lines += ["", "## Focus query table", ""]
    for v in variants:
        lines += [f"### {v}", "", "| query | bucket | qual | hits (collection:key) |", "|---|---|---|---|"]
        for row in report["variants"][v]["focus"]:
            if row["id"].startswith("focus-style-"):
                continue
            hit_s = ", ".join(f"{h['collection']}:{h['key']}" for h in row["hits"][:6]) or "∅"
            lines.append(f"| `{row['query']}` | {row['bucket']} | **{row['qual']}** | {hit_s} |")
        lines.append("")
        # style extras brief
        styles = [r for r in report["variants"][v]["focus"] if r["id"].startswith("focus-style-")]
        if styles:
            lines.append("I078/I036/I210-style iteration queries:")
            for row in styles:
                hit_s = ", ".join(f"{h['collection']}:{h['key']}" for h in row["hits"][:4]) or "∅"
                lines.append(f"- `{row['query']}` → {row['qual']} | {hit_s}")
            lines.append("")

    lines += ["", "## Solves old problems? (MISS→OK flips)", "", "| query | V0 | V1 | V2 | V3 |", "|---|---|---|---|---|"]
    primary_ids = {
        "关爱客户": "focus-guanai",
        "有钱": "focus-youqian",
        "有钱客户": "focus-youqian-kehu",
        "金凤": "focus-jinfeng",
        "盘客": "focus-panke",
        "去盘客": "focus-qupanke",
    }
    for q, fid in primary_ids.items():
        cells = []
        for v in variants:
            m = {r["id"]: r["qual"] for r in report["variants"][v]["focus"]}
            cells.append(m.get(fid, "-"))
        lines.append(f"| `{q}` | " + " | ".join(cells) + " |")

    lines += [
        "",
        "## New problems / false accepts",
        "",
        "| variant | weather | hobby | latin A | 陈金秀 | stress irr/unsup false-accept proxy |",
        "|---|---|---|---|---|---|",
    ]
    for v in variants:
        m = {r["query"]: r for r in report["variants"][v]["focus"]}
        fa = report["variants"][v]["stress_summary"].get("false_accept_proxy") or {}
        lines.append(
            f"| {v} | {m.get('天气怎么样',{}).get('qual')} | {m.get('客户平时有什么兴趣爱好',{}).get('qual')} | "
            f"{m.get('A',{}).get('qual')} | {m.get('陈金秀',{}).get('qual')} | {fa} |"
        )

    lines += ["", "## Stress buckets", ""]
    for v in variants:
        lines += [f"### {v}", "", "| bucket | n | with_hits | empty |", "|---|---:|---:|---:|"]
        for b, c in report["variants"][v]["stress_summary"]["buckets"].items():
            lines.append(f"| {b} | {c.get('n',0)} | {c.get('with_hits',0)} | {c.get('empty',0)} |")
        lines.append("")

    lines += [
        "",
        "## Interpretation",
        "",
    ]

    # auto interpretation from numbers
    def qual(v, q):
        for r in report["variants"][v]["focus"]:
            if r["query"] == q:
                return r["qual"]
        return "?"

    helped = []
    for q in ("关爱客户", "有钱", "金凤", "盘客"):
        if qual("V0", q) == "MISS" and any(qual(v, q) == "OK" for v in ("V1", "V2", "V3")):
            who = [v for v in ("V1", "V2", "V3") if qual(v, q) == "OK"]
            helped.append(f"{q}: V0 MISS → OK via {who}")
    lines.append("### Did V1/V2/V3 help?")
    if helped:
        for h in helped:
            lines.append(f"- {h}")
    else:
        lines.append("- No clear MISS→OK flips on the four Auth-OFF nouns.")

    lines.append("")
    lines.append("### What broke / new risks?")
    for v in variants:
        weather = qual(v, "天气怎么样")
        hobby = qual(v, "客户平时有什么兴趣爱好")
        a = qual(v, "A")
        name = qual(v, "陈金秀")
        d = report["variants"][v]["labeled_metrics"]["development"]
        h = report["variants"][v]["labeled_metrics"]["holdout"]
        lines.append(
            f"- {v}: weather={weather}, hobby={hobby}, A={a}, 陈金秀={name}; "
            f"dev irr_rej={d.get('irrelevant_rejection_rate')}, holdout irr_rej={h.get('irrelevant_rejection_rate')}, "
            f"holdout top8={h.get('top8_recall_rate')}"
        )

    lines += [
        "",
        "### Generalization",
        "",
        "- Dual thresholds (top8≥0.85 AND irr_rej=1.0 on **both** splits) still required for formal selection — this sim does not claim selection.",
        "- Holdout paraphrases remain the hard failure mode for field lexical strategies.",
        "- Latin barewords / person names / enum nouns need multi-collection exact routing (V2), not field char-heuristic alone.",
        "",
        f"Raw JSON: `{OUT_JSON}`",
        "",
    ]
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")

    # append issues-log
    brief = []
    brief.append("\n## Empirical key-index sim (2026-08-13 UTC+8)\n")
    brief.append("Offline V0–V3 measurement (not selected/promoted).\n")
    for v in variants:
        d = report["variants"][v]["labeled_metrics"]["development"]
        h = report["variants"][v]["labeled_metrics"]["holdout"]
        s = report["variants"][v]["stress_summary"]
        flips = report["variants"][v]["focus_flips"]
        brief.append(
            f"- **{v}**: dev top8={d.get('top8_recall_rate'):.3f}/irr={d.get('irrelevant_rejection_rate'):.3f}; "
            f"holdout top8={h.get('top8_recall_rate'):.3f}/irr={h.get('irrelevant_rejection_rate'):.3f}; "
            f"stress empty={s['empty_rate']:.3f}; flips={flips}\n"
        )
    brief.append(f"- Artifacts: `{OUT_MD}`, `{OUT_JSON}`\n")
    with ISSUES.open("a", encoding="utf-8") as f:
        f.writelines(brief)

    print(
        json.dumps(
            {
                "out_json": str(OUT_JSON),
                "out_md": str(OUT_MD),
                "smoke": smoke,
                "focus_flips": {v: report["variants"][v]["focus_flips"] for v in variants},
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
